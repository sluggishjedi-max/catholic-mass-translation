from __future__ import annotations

import argparse
import html
import io
import json
import re
import shutil
import sys
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup
from PIL import Image


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


ROOT = Path(__file__).resolve().parents[1]
HYMN_DATA = ROOT / "JS file" / "hymn_data.js"
ASSETS = ROOT / "assets" / "hymns"
TMP = ROOT / "tmp" / "online-hymn-import"

US_UPDATE_2025_URL = "https://www.icrmusic.org/en-us/14/page/331"
US_UPDATE_2026_URL = "https://www.icrmusic.org/en-us/14/page/503"
US_SEARCH_URL = "https://www.icrmusic.org/en-us/14/cluster/search"
US_CONTENTS_URL = (
    "https://d5uh4t1x1kne0.cloudfront.net/breakingbread/"
    "Breaking%20Bread%202026%20Contents%20Excel.xlsx"
)
US_BOOK = "Breaking Bread Digital Music Library"
US_ATTRIBUTION = "ⓒ Breaking Bread (2026 edition), OCP"
US_ID_PREFIX = "en-breaking-bread-2026-"
US_ASSET_DIR = ASSETS / "en-breaking-bread-2026"
US_EXPECTED = 866
US_UPDATE_EXPECTED = {2025: 29, 2026: 21}

JP_PRIMARY_INDEX = "http://tenreiseika.romaaeterna.jp/antiphon/index.html"
JP_SUPPLEMENT_INDEX = "http://hosanna.romaaeterna.jp/tenreiseika/tseika.html"
JP_BOOK = "典礼聖歌"
JP_ID_PREFIX = "jp-tenrei-"
JP_ASSET_DIR = ASSETS / "jp-tenrei"
JP_PRIMARY_EXPECTED = 202
JP_SUPPLEMENT_EXPECTED = 47
JP_EXPECTED = JP_PRIMARY_EXPECTED + JP_SUPPLEMENT_EXPECTED

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140 Safari/537.36"
    )
}


@dataclass(frozen=True)
class IndexLink:
    stem: str
    label: str
    url: str
    source: str


def fetch(url: str, *, attempts: int = 3, timeout: int = 30) -> requests.Response:
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            response = requests.get(url, headers=HEADERS, timeout=timeout)
            response.raise_for_status()
            return response
        except Exception as exc:  # pragma: no cover - exercised only on network failure
            last_error = exc
            if attempt < attempts:
                time.sleep(0.6 * attempt)
    raise RuntimeError(f"Failed to download {url}: {last_error}")


def decode_html(content: bytes, fallback: str = "utf-8") -> str:
    head = content[:8192].decode("ascii", errors="ignore")
    match = re.search(
        r"(?:charset\s*=\s*|encoding\s*=\s*)[\"']?([A-Za-z0-9._-]+)",
        head,
        flags=re.IGNORECASE,
    )
    candidates = [match.group(1) if match else "", fallback, "utf-8", "cp932"]
    for encoding in dict.fromkeys(candidate for candidate in candidates if candidate):
        try:
            return content.decode(encoding)
        except (LookupError, UnicodeDecodeError):
            continue
    return content.decode(fallback, errors="replace")


def clean_text(value: str) -> str:
    return " ".join(html.unescape(value or "").replace("\u3000", " ").split())


def html_text(fragment: str) -> str:
    return clean_text(BeautifulSoup(fragment or "", "html.parser").get_text(" ", strip=True))


def normalized(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(character for character in value if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def slugify(value: str, fallback: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(character for character in value if not unicodedata.combining(character))
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return value or fallback


def unique(values: list[str]) -> list[str]:
    return list(dict.fromkeys(clean_text(value) for value in values if clean_text(value)))


def relative(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def find_hymn_array(text: str) -> list[dict]:
    marker = "const hymnData"
    start = text.index(marker)
    array_start = text.index("[", start)
    depth = 0
    in_string = False
    escaped = False
    for index in range(array_start, len(text)):
        character = text[index]
        if in_string:
            if escaped:
                escaped = False
            elif character == "\\":
                escaped = True
            elif character == '"':
                in_string = False
            continue
        if character == '"':
            in_string = True
        elif character == "[":
            depth += 1
        elif character == "]":
            depth -= 1
            if depth == 0:
                return json.loads(text[array_start : index + 1])
    raise ValueError("Could not find the hymn data array")


def load_hymn_data() -> list[dict]:
    return find_hymn_array(HYMN_DATA.read_text(encoding="utf-8"))


def serialize_hymn_data(entries: list[dict]) -> str:
    body = json.dumps(entries, ensure_ascii=False, indent=2).replace("\u2028", "\\u2028").replace(
        "\u2029", "\\u2029"
    )
    return "\n".join(
        [
            "// Hymn index and compressed per-song score images for V22.1.",
            "// Generated by tools/generate_hymn_assets.py; manual entries can be updated with tools/hymn-data-entry-tool.js.",
            "(function attachHymnData(global) {",
            f"  const hymnData = {body};",
            "  global.hymnData = Array.isArray(global.hymnData) && global.hymnData.length",
            "    ? global.hymnData",
            "    : hymnData;",
            "  global.ordoHymnData = hymnData;",
            "})(typeof window !== 'undefined' ? window : globalThis);",
            "",
        ]
    )


def save_hymn_data(entries: list[dict]) -> Path | None:
    ids = [clean_text(str(entry.get("id", ""))) for entry in entries]
    missing = [index for index, value in enumerate(ids) if not value]
    duplicates = sorted({value for value in ids if ids.count(value) > 1})
    if missing:
        raise RuntimeError(f"Hymn entries without IDs: {missing[:10]}")
    if duplicates:
        raise RuntimeError(f"Duplicate hymn IDs: {duplicates[:10]}")

    output = serialize_hymn_data(entries)
    current = HYMN_DATA.read_text(encoding="utf-8")
    if output == current:
        return None

    TMP.mkdir(parents=True, exist_ok=True)
    backup = TMP / f"hymn_data.backup-{int(time.time())}.js"
    shutil.copyfile(HYMN_DATA, backup)
    temporary = HYMN_DATA.with_suffix(HYMN_DATA.suffix + ".online-import.tmp")
    temporary.write_text(output, encoding="utf-8", newline="\n")
    temporary.replace(HYMN_DATA)
    return backup


def image_to_webp(content: bytes, output: Path) -> tuple[int, int]:
    with Image.open(io.BytesIO(content)) as image:
        image.load()
        if image.width < 100 or image.height < 80:
            raise RuntimeError(f"Unexpectedly small score image: {image.width}x{image.height}")
        if image.mode in {"RGBA", "LA"} or "transparency" in image.info:
            rgba = image.convert("RGBA")
            background = Image.new("RGBA", rgba.size, "white")
            background.alpha_composite(rgba)
            rgb = background.convert("RGB")
        else:
            rgb = image.convert("RGB")
        output.parent.mkdir(parents=True, exist_ok=True)
        rgb.save(output, "WEBP", lossless=True, method=6)
        return image.width, image.height


def breaking_bread_rows(content: bytes) -> list[dict]:
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    values = list(sheet.iter_rows(values_only=True))
    headers = [clean_text(str(value or "")) for value in values[0]]
    return [dict(zip(headers, row)) for row in values[1:] if any(value not in {None, ""} for value in row)]


def parse_us_catalog(source_html: str) -> list[dict]:
    soup = BeautifulSoup(source_html, "html.parser")
    links: dict[str, dict] = {}
    for anchor in soup.select("a[href]"):
        href = clean_text(anchor.get("href", ""))
        match = re.fullmatch(r"(?:/en-us)?/14/cluster/(\d+)", href)
        if not match:
            continue
        row = anchor.find_parent("tr")
        cells = row.find_all("td") if row else []
        if len(cells) < 2:
            continue
        number_text = clean_text(cells[1].get_text(" ", strip=True))
        if not number_text.isdigit():
            raise RuntimeError(f"Unexpected Breaking Bread number {number_text!r}")
        title = clean_text(anchor.get_text(" ", strip=True))
        title_and_credit = clean_text(cells[0].get_text(" ", strip=True))
        credit = title_and_credit[len(title):].strip() if title_and_credit.startswith(title) else ""
        cluster = match.group(1)
        links[cluster] = {
            "cluster": cluster,
            "url": f"https://www.icrmusic.org/en-us/14/cluster/{cluster}",
            "number": int(number_text),
            "title": title,
            "credit": credit,
        }
    output = list(links.values())
    if len(output) != US_EXPECTED:
        raise RuntimeError(f"Expected {US_EXPECTED} Breaking Bread catalog links, found {len(output)}")
    numbers = [link["number"] for link in output]
    if len(numbers) != len(set(numbers)):
        raise RuntimeError("Breaking Bread catalog contains duplicate hymn numbers")
    return output


def parse_us_update_statuses(source_html: str, year: int) -> dict[str, str]:
    soup = BeautifulSoup(source_html, "html.parser")
    statuses: dict[str, str] = {}
    for anchor in soup.select("a[href]"):
        href = clean_text(anchor.get("href", ""))
        match = re.fullmatch(r"(?:/en-us)?/14/cluster/(\d+)", href)
        if not match:
            continue
        heading = anchor.find_previous(["h2", "h3"])
        heading_text = normalized(heading.get_text(" ", strip=True) if heading else "")
        if "songs added" in heading_text:
            status = "added"
        elif "songs being changed" in heading_text or "songs changed" in heading_text:
            status = "changed"
        else:
            continue
        statuses[match.group(1)] = status
    expected = US_UPDATE_EXPECTED[year]
    if len(statuses) != expected:
        raise RuntimeError(f"Expected {expected} Breaking Bread {year} updates, found {len(statuses)}")
    return statuses


def breaking_bread_rows_by_number(rows: list[dict]) -> dict[int, dict]:
    output = {}
    for row in rows:
        value = row.get("Song #")
        if isinstance(value, (int, float)) and int(value) == value:
            output[int(value)] = row
    return output


def parse_icr_preview(page_html: str, page_url: str) -> tuple[str, str]:
    soup = BeautifulSoup(page_html, "html.parser")
    candidates: list[tuple[int, str, str]] = []
    for asset in soup.select(".sidebar_cluster_asset[sku_id]"):
        title_node = asset.select_one(".asset_container_title")
        format_node = asset.select_one(".right")
        asset_title = clean_text(title_node.get_text(" ", strip=True) if title_node else "")
        asset_format = clean_text(format_node.get_text(" ", strip=True) if format_node else "")
        sku = clean_text(asset.get("sku_id", ""))
        image = soup.select_one(f"#img-{sku}")
        preview = image.get("url", "") if image else ""
        if not sku or not preview:
            continue
        title_key = normalized(asset_title)
        if title_key == "congregational sheet music":
            priority = 0
        elif "lead sheet" in title_key:
            priority = 1
        elif "keyboard accompaniment" in title_key:
            priority = 2
        elif "guitar accompaniment" in title_key:
            priority = 3
        elif asset_format in {"PDF", "GIF"}:
            priority = 4
        else:
            continue
        candidates.append((priority, sku, urljoin(page_url, preview)))
    if candidates:
        _priority, sku, preview = min(candidates)
        return sku, preview
    raise RuntimeError(f"No score preview found at {page_url}")


def import_us_hymns() -> tuple[list[dict], dict]:
    print("[US] Fetching the complete Breaking Bread catalog...", flush=True)
    catalog_response = fetch(US_SEARCH_URL)
    catalog_html = decode_html(catalog_response.content, "utf-8")
    update_2025_response = fetch(US_UPDATE_2025_URL)
    update_2025_html = decode_html(update_2025_response.content, "utf-8")
    update_2026_response = fetch(US_UPDATE_2026_URL)
    update_2026_html = decode_html(update_2026_response.content, "utf-8")
    contents_response = fetch(US_CONTENTS_URL)
    rows_by_number = breaking_bread_rows_by_number(breaking_bread_rows(contents_response.content))
    links = parse_us_catalog(catalog_html)
    update_statuses = {
        2025: parse_us_update_statuses(update_2025_html, 2025),
        2026: parse_us_update_statuses(update_2026_html, 2026),
    }

    TMP.mkdir(parents=True, exist_ok=True)
    (TMP / "breaking-bread-catalog.html").write_text(catalog_html, encoding="utf-8")
    (TMP / "breaking-bread-2025-update.html").write_text(update_2025_html, encoding="utf-8")
    (TMP / "breaking-bread-2026-update.html").write_text(update_2026_html, encoding="utf-8")
    (TMP / "breaking-bread-2026-contents.xlsx").write_bytes(contents_response.content)
    cluster_cache_dir = TMP / "icr-clusters"
    cluster_cache_dir.mkdir(parents=True, exist_ok=True)

    def build(link: dict) -> dict:
        row = rows_by_number.get(link["number"], {})
        cache_path = cluster_cache_dir / f"{link['cluster']}.html"
        if cache_path.is_file() and cache_path.stat().st_size > 0:
            page_html = cache_path.read_text(encoding="utf-8")
        else:
            page_response = fetch(link["url"])
            page_html = decode_html(page_response.content, "utf-8")
            cache_path.write_text(page_html, encoding="utf-8")
        sku, preview_url = parse_icr_preview(page_html, link["url"])

        number_value = link["number"]
        number = f"{number_value:03d}"
        title = link["title"]
        credit = clean_text(link["credit"])
        stem = f"{number}-{slugify(title, link['cluster'])}-01.webp"
        output = US_ASSET_DIR / stem
        if output.is_file() and output.stat().st_size > 0:
            with Image.open(output) as cached_image:
                width, height = cached_image.size
            original_size = output.stat().st_size
        else:
            image_response = fetch(preview_url)
            original_size = len(image_response.content)
            width, height = image_to_webp(image_response.content, output)

        book_section = clean_text(str(row.get("Book Section") or ""))
        liturgical_moment = clean_text(str(row.get("Liturgical Moment") or ""))
        source_updates = []
        for year in sorted(update_statuses):
            status = update_statuses[year].get(link["cluster"])
            if status:
                source_updates.append(f"{status.title()} in {year}")
        tags = unique([US_BOOK, book_section, liturgical_moment, *source_updates])
        aliases = unique(
            [
                title,
                credit,
                f"Breaking Bread {number_value}",
                f"Breaking Bread 2026 {number_value}",
                f"BB 2026 {number_value}",
                f"BB{number}",
            ]
        )
        return {
            "id": f"{US_ID_PREFIX}{number}",
            "country": "EN",
            "language": "EN",
            "number": number,
            "title": title,
            "displayTitle": f"{number}. {title}",
            "book": US_BOOK,
            "tags": tags,
            "category": " / ".join(tags),
            "lyrics": "",
            "translations": {},
            "firstLine": "",
            "composer": credit,
            "lyricist": "",
            "sourceFormat": "PNG preview of PDF",
            "originalFileName": f"{sku}.png",
            "originalFileSize": original_size,
            "originalFileAvailable": True,
            "searchAliases": aliases,
            "scoreImages": [{"src": relative(output), "label": "1"}],
            "scoreNote": "Public low-resolution ICR/OCP congregational sheet-music preview.",
            "copyright": US_ATTRIBUTION,
            "sourceUrl": link["url"],
            "sourcePreviewUrl": preview_url,
            "sourceUpdates": source_updates,
            "sourceImageDimensions": f"{width}x{height}",
        }

    entries: list[dict] = []
    with ThreadPoolExecutor(max_workers=12) as executor:
        futures = [executor.submit(build, link) for link in links]
        for future in as_completed(futures):
            entry = future.result()
            entries.append(entry)
            count = len(entries)
            if count % 25 == 0 or count == US_EXPECTED:
                print(f"[US] {count:03d}/{US_EXPECTED}: {entry['displayTitle']}", flush=True)
    entries.sort(key=lambda entry: (int(entry["number"]), entry["title"]))
    if len(entries) != US_EXPECTED or any(not entry["scoreImages"] for entry in entries):
        raise RuntimeError("Breaking Bread catalog import did not produce every expected score preview")
    return entries, {
        "source": US_SEARCH_URL,
        "entries": len(entries),
        "updates": {
            str(year): {
                "source": US_UPDATE_2025_URL if year == 2025 else US_UPDATE_2026_URL,
                "added": sum(status == "added" for status in statuses.values()),
                "changed": sum(status == "changed" for status in statuses.values()),
            }
            for year, statuses in update_statuses.items()
        },
        "scoreImages": sum(len(entry["scoreImages"]) for entry in entries),
    }


def parse_jp_index(content: bytes, index_url: str, source: str, fallback: str) -> list[IndexLink]:
    soup = BeautifulSoup(decode_html(content, fallback), "html.parser")
    links: dict[str, IndexLink] = {}
    for anchor in soup.find_all("a", href=True):
        url = urljoin(index_url, anchor["href"])
        path = urlparse(url).path
        if not path.lower().endswith(".html"):
            continue
        stem = Path(path).stem.lower()
        if not re.match(r"(?:ten|tsei)\d+", stem):
            continue
        label = clean_text(anchor.get_text(" ", strip=True))
        if label and label != "♪":
            links.setdefault(url, IndexLink(stem=stem, label=label, url=url, source=source))
    return list(links.values())


def jp_categories(text: str, antiphon_text: str, source_url: str) -> list[str]:
    tags = [JP_BOOK]
    combined = f"{text} {antiphon_text}"
    if "/alleluiah/" in source_url or "詠唱" in combined:
        tags.append("詠唱")
    if "答唱" in antiphon_text or "詩編" in combined:
        tags.append("答唱詩編")
    if "交唱" in antiphon_text:
        tags.append("交唱")
    mapping = [
        ("待降", "待降節"),
        ("降誕", "降誕節"),
        ("四旬", "四旬節"),
        ("聖週間", "聖週間"),
        ("復活", "復活節"),
        ("年間", "年間"),
        ("聖霊", "聖霊"),
        ("聖母", "聖母"),
        ("マリア", "聖母"),
        ("葬儀", "葬儀"),
        ("入祭", "入祭"),
        ("閉祭", "閉祭"),
        ("朝の祈り", "朝の祈り"),
        ("晩の祈り", "晩の祈り"),
        ("洗礼", "洗礼"),
        ("回心", "回心"),
        ("感謝", "感謝"),
        ("信頼", "信頼"),
        ("希望", "希望"),
    ]
    for needle, tag in mapping:
        if needle in combined:
            tags.append(tag)
    return unique(tags)


def parse_credits(metadata: str) -> tuple[str, str]:
    match = re.search(r"[（(]\s*詞\s*[：:]\s*(.*?)\s*/\s*曲\s*[：:]\s*(.*?)[）)]", metadata)
    if not match:
        return "", ""
    return clean_text(match.group(2)), clean_text(match.group(1))


def parse_scripture(metadata_lines: list[str]) -> str:
    scripture_names = "詩編|ルカ|マタイ|マルコ|ヨハネ|コリント|ローマ|イザヤ|エフェソ|フィリピ|黙示録"
    for line in metadata_lines[1:]:
        if re.search(scripture_names, line) and re.search(r"\d", line):
            return clean_text(line)
    return ""


def jp_image_urls(soup: BeautifulSoup, page_url: str) -> list[str]:
    urls = []
    for image in soup.find_all("img", src=True):
        url = urljoin(page_url, image["src"])
        name = Path(urlparse(url).path).name.lower()
        if not re.search(r"\.(?:gif|png|jpe?g)$", name):
            continue
        if name in {"mp3.gif"} or "icon" in name or "button" in name:
            continue
        urls.append(url)
    return list(dict.fromkeys(urls))


def japanese_page_entry(link: IndexLink) -> dict:
    response = fetch(link.url)
    fallback = "cp932" if link.source == "tenreiseika-antiphon" else "utf-8"
    page_html = decode_html(response.content, fallback)
    soup = BeautifulSoup(page_html, "html.parser")

    h2 = clean_text(soup.h2.get_text(" ", strip=True) if soup.h2 else "")
    title = re.sub(r"^\d+\s*", "", h2).strip() or link.label
    number_match = re.match(r"(?:ten|tsei)(\d+)", link.stem)
    if not number_match:
        raise RuntimeError(f"Could not determine a hymn number from {link.stem}")
    number = f"{int(number_match.group(1)):03d}"

    metadata_lines: list[str] = []
    if soup.h4:
        metadata_lines = unique(soup.h4.get_text("\n", strip=True).splitlines())
    metadata = " ".join(metadata_lines)
    composer, lyricist = parse_credits(metadata)
    scripture = parse_scripture(metadata_lines)

    antiphons = unique(node.get_text("\n", strip=True) for node in soup.select(".ant"))
    lyrics = "\n\n".join(antiphons)
    if not lyrics and link.source == "hosanna-supplement":
        paragraphs = []
        for paragraph in soup.select(".box p"):
            classes = set(paragraph.get("class", []))
            value = clean_text(paragraph.get_text(" ", strip=True))
            if not value or classes.intersection({"pb", "info1", "info2"}) or value.startswith("Update:"):
                continue
            paragraphs.append(value)
        lyrics = "\n\n".join(unique(paragraphs))

    all_page_text = clean_text(soup.get_text(" ", strip=True))
    tags = jp_categories(f"{metadata} {all_page_text}", lyrics, link.url)
    image_urls = jp_image_urls(soup, link.url)
    image_records = []
    original_names = []
    original_size = 0
    for index, image_url in enumerate(image_urls, start=1):
        image_response = fetch(image_url)
        output = JP_ASSET_DIR / f"{link.stem}-{index:02d}.webp"
        image_to_webp(image_response.content, output)
        image_records.append({"src": relative(output), "label": str(index)})
        original_names.append(Path(urlparse(image_url).path).name)
        original_size += len(image_response.content)
    if len(image_records) > 1:
        for index, record in enumerate(image_records, start=1):
            record["label"] = f"{index}/{len(image_records)}"

    aliases = unique(
        [
            title,
            f"典礼聖歌 {int(number)}",
            f"典礼聖歌 {number}",
            f"典礼聖歌 {int(number)}番",
            link.stem,
            scripture,
        ]
    )
    first_line = ""
    for line in lyrics.splitlines():
        candidate = re.sub(r"^【[^】]+】\s*", "", clean_text(line))
        if candidate:
            first_line = candidate
            break

    return {
        "id": f"{JP_ID_PREFIX}{link.stem}",
        "country": "JP",
        "language": "JP",
        "number": number,
        "title": title,
        "displayTitle": f"{number}. {title}",
        "book": JP_BOOK,
        "tags": tags,
        "category": " / ".join(tags),
        "lyrics": lyrics,
        "translations": {},
        "firstLine": first_line,
        "composer": composer,
        "lyricist": lyricist,
        "scripture": scripture,
        "sourceFormat": "GIF" if image_records else "HTML",
        "originalFileName": "; ".join(original_names),
        "originalFileSize": original_size,
        "originalFileAvailable": bool(image_records),
        "searchAliases": aliases,
        "scoreImages": image_records,
        "scoreNote": "" if image_records else "Source page publishes chant text/audio only; no score image is available.",
        "copyright": "典礼聖歌 / JASRAC 管理曲. Used with permission confirmed by the project owner.",
        "sourceUrl": link.url,
        "sourceCollection": link.source,
    }


def import_japanese_hymns() -> tuple[list[dict], dict]:
    print("[JP] Fetching the Japanese hymn indexes...", flush=True)
    primary_response = fetch(JP_PRIMARY_INDEX)
    supplement_response = fetch(JP_SUPPLEMENT_INDEX)
    primary_links = parse_jp_index(
        primary_response.content, JP_PRIMARY_INDEX, "tenreiseika-antiphon", "cp932"
    )
    supplement_links = parse_jp_index(
        supplement_response.content, JP_SUPPLEMENT_INDEX, "hosanna-supplement", "utf-8"
    )
    if len(primary_links) != JP_PRIMARY_EXPECTED:
        raise RuntimeError(f"Expected {JP_PRIMARY_EXPECTED} primary Japanese pages, found {len(primary_links)}")

    primary_keys = {re.sub(r"^(?:ten|tsei)", "", link.stem) for link in primary_links}
    supplement_links = [
        link
        for link in supplement_links
        if re.sub(r"^(?:ten|tsei)", "", link.stem) not in primary_keys
    ]
    if len(supplement_links) != JP_SUPPLEMENT_EXPECTED:
        raise RuntimeError(
            f"Expected {JP_SUPPLEMENT_EXPECTED} unique Japanese supplement pages, found {len(supplement_links)}"
        )

    TMP.mkdir(parents=True, exist_ok=True)
    (TMP / "jp-primary-index.html").write_text(
        decode_html(primary_response.content, "cp932"), encoding="utf-8"
    )
    (TMP / "jp-supplement-index.html").write_text(
        decode_html(supplement_response.content, "utf-8"), encoding="utf-8"
    )

    entries: list[dict] = []
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = [executor.submit(japanese_page_entry, link) for link in primary_links + supplement_links]
        for future in as_completed(futures):
            entry = future.result()
            entries.append(entry)
            count = len(entries)
            if count == 1 or count % 10 == 0 or count == JP_EXPECTED:
                print(f"[JP] {count:03d}/{JP_EXPECTED}: {entry['displayTitle']}", flush=True)
    entries.sort(key=lambda entry: (int(entry["number"]), entry["id"]))
    if len(entries) != JP_EXPECTED:
        raise RuntimeError(f"Expected {JP_EXPECTED} Japanese entries, built {len(entries)}")

    return entries, {
        "sources": [JP_PRIMARY_INDEX, JP_SUPPLEMENT_INDEX],
        "entries": len(entries),
        "primaryEntries": sum(entry["sourceCollection"] == "tenreiseika-antiphon" for entry in entries),
        "supplementEntries": sum(entry["sourceCollection"] == "hosanna-supplement" for entry in entries),
        "entriesWithScores": sum(bool(entry["scoreImages"]) for entry in entries),
        "scoreImages": sum(len(entry["scoreImages"]) for entry in entries),
    }


def validate_current(entries: list[dict]) -> dict:
    ids = [clean_text(str(entry.get("id", ""))) for entry in entries]
    duplicate_ids = sorted({value for value in ids if ids.count(value) > 1})
    if duplicate_ids:
        raise RuntimeError(f"Duplicate hymn IDs: {duplicate_ids[:10]}")

    us_entries = [entry for entry in entries if str(entry.get("id", "")).startswith(US_ID_PREFIX)]
    jp_entries = [entry for entry in entries if str(entry.get("id", "")).startswith(JP_ID_PREFIX)]
    if len(us_entries) != US_EXPECTED:
        raise RuntimeError(f"Expected {US_EXPECTED} imported US entries, found {len(us_entries)}")
    if len(jp_entries) != JP_EXPECTED:
        raise RuntimeError(f"Expected {JP_EXPECTED} imported Japanese entries, found {len(jp_entries)}")

    missing_assets = []
    empty_assets = []
    for entry in us_entries + jp_entries:
        for score in entry.get("scoreImages", []):
            path = ROOT / str(score.get("src", ""))
            if not path.is_file():
                missing_assets.append(str(path))
            elif path.stat().st_size <= 0:
                empty_assets.append(str(path))
    if missing_assets or empty_assets:
        raise RuntimeError(
            f"Missing score assets: {missing_assets[:5]}; empty score assets: {empty_assets[:5]}"
        )
    if any(not entry.get("scoreImages") for entry in us_entries):
        raise RuntimeError("Every imported US entry must include a score preview")

    return {
        "totalHymns": len(entries),
        "usEntries": len(us_entries),
        "usScoreImages": sum(len(entry.get("scoreImages", [])) for entry in us_entries),
        "jpEntries": len(jp_entries),
        "jpEntriesWithScores": sum(bool(entry.get("scoreImages")) for entry in jp_entries),
        "jpScoreImages": sum(len(entry.get("scoreImages", [])) for entry in jp_entries),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import public US and Japanese hymn resources")
    parser.add_argument("--check", action="store_true", help="Validate the already imported data and assets")
    args = parser.parse_args()

    existing = load_hymn_data()
    if args.check:
        print(json.dumps(validate_current(existing), ensure_ascii=False, indent=2))
        return

    us_entries, us_report = import_us_hymns()
    jp_entries, jp_report = import_japanese_hymns()
    retained = [
        entry
        for entry in existing
        if not str(entry.get("id", "")).startswith((US_ID_PREFIX, JP_ID_PREFIX))
    ]
    combined = retained + us_entries + jp_entries
    backup = save_hymn_data(combined)
    validation = validate_current(combined)

    report = {
        "breakingBreadCatalog": us_report,
        "japanese": jp_report,
        "validation": validation,
        "hymnDataBackup": str(backup) if backup else None,
        "stMichael4thEdition": {
            "source": "https://www.stmichaelhymnal.com/download-hymns/4th-edition-downloads/",
            "status": "Login required for PDF downloads; no protected files were imported.",
        },
    }
    TMP.mkdir(parents=True, exist_ok=True)
    (TMP / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
